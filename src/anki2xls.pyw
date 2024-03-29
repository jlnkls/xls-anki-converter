"""ANKI (TXT) 2 XLS

Python script that:
- Takes as a first input an Anki-ready TXT file that contains a table separated by tabs:
--> the first 5 rows are metadata:
#separator:tab
#html:false
#guid column:1
#notetype column:2
#tags column:3

--> the rest of the rows are actual table data separated by tabs, like this:
KR-VG&&]+fv((   Basic (type in the answer)  ser 이다  

- Takes as a second input an XLS document

- Reads the metadata in the TXT (rows 1 to 5) in a metadata DataFrame

- Reads the rest of the data in the TXT (rows 6 onwards) in a data DataFrame

- Copies the rows of the metadata DataFrame to rows 2 to 6 in the XLS file (that is, it substitutes the original content of the XLSM file from rows 2 to 6)

- Copies the data DataFrame from rows 8 to the end of the XLS (adding more rows in the XLS DataFrame if the TXT file has more rows than the XLS DataFrame) in the following fashion:
--> It copies column 1 of the TXT in column 1 of the XLSM (from row 8 to the end of the XLSM) (Anki GUID),
--> It copies column 2 of the TXT in column 5 of the XLSM (from row 8 to the end of the XLSM) (Anki Notetype),
--> It copies column 3 of the TXT in column 3 of the XLSM (from row 8 to the end of the XLSM) (source language)
--> It copies column 4 of the TXT in column 2 of the XLSM (from row 8 to the end of the XLSM) (language being learned)
--> Removes all the contents from column 4 of the XLSM (from row 8 to the end of the XLSM) (Anki Tags)
---> (that is, it substitutes the original content of the XLS file from row 8 to the end)

- Changes the font of every cell from row 9 onwards in the XLS to font size "15"

- Keeps the original formatting of the XLS file, the colors, the visible print area, the freeze panes option, and the VBA code (if it is .xlsm)
--> Usees the openpyxl library

- Syntax:
./anki2xls.pyw "$$$TWO_OR_THREE_LETTER_IDENTIFIER_OF_LANGUAGE$$$"

Author: jlnkls
"""


import pandas as pd
import sys
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment


def anki2xls(path, anki_txt_filename, vocab_list_name, notetype):
    ''' Update the vocabulary spreadsheet with the Anki-output TXT file '''

    # Read the TXT file
    txt_metadata = pd.read_csv(path + "Anki Export/" + anki_txt_filename, sep='\t', nrows=5, header=None, index_col=False)
    txt_data = pd.read_csv(path + "Anki Export/" + anki_txt_filename, sep='\t', skiprows=5, header=None, index_col=False)
    
    # Read the XLS file
    xl = openpyxl.load_workbook(path + vocab_list_name + ".xlsm", keep_vba=True)
    sheet = xl.active

    # Check and modify if a cell in txt_metadata contains "#tags column:" or "#notetype column:"
    for r_idx, row in txt_metadata.iterrows():
        for c_idx, value in enumerate(row):
            if isinstance(value, str):
                if "#tags column:" in value:
                    tag_value = value.split(":")[-1].strip()
                    if tag_value.isdigit():  # Check if the value after ":" is a number
                        txt_metadata.iat[r_idx, c_idx] = value.replace(tag_value, "4") # Make sure the TAG column is marked as the 4th column
                elif "#notetype column:" in value:
                    notetype_value = value.split(":")[-1].strip()
                    txt_metadata.iat[r_idx, c_idx] = value.replace(notetype_value, "5") # Make sure the Notetype column is marked as the 5th column

    # Copy metadata from rows 2 to 5 in the XLS file
    for r_idx, row in txt_metadata.iterrows():
        for c_idx, value in enumerate(row):
            sheet.cell(row=r_idx + 2, column=c_idx + 1, value=value)

    # Clear all content from row 9 onwards in all columns
    for row in sheet.iter_rows(min_row=9):
        for cell in row:
            cell.value = None

    # Copy data to row 9 and onwards in the XLS file
    for r_idx, row in txt_data.iterrows():
        sheet.cell(row=r_idx + 9, column=1, value=row[0])  # Copy column 1 of TXT to column 1 of XLS (Anki GUID)
        sheet.cell(row=r_idx + 9, column=5, value=row[1])  # Copy column 2 of TXT to column 5 of XLS (Notetype)
        sheet.cell(row=r_idx + 9, column=3, value=row[2])  # Copy column 3 of TXT to column 3 of XLS (source language)
        sheet.cell(row=r_idx + 9, column=2, value=row[3])  # Copy column 4 of TXT to column 2 of XLS (language being learned)
        sheet.cell(row=r_idx + 9, column=4, value=row[4])  # Copy column 5 of TXT to column 4 of XLS (Anki Tags)


    # Check and modify cells starting with "=" (escape them as Excel treats them otherwise as formulas)
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = ' ' + cell.value

    # Define the font style as white (color) for row one
    font_style_white = Font(color="FFFFFF", name='Segoe UI', size=15, bold=True)  # White color

    # Change the font color to white in row one
    for cell in sheet[1]:
        cell.font = font_style_white

    # Define the font style as bold for row 8
    font_style_bold = Font(name='Segoe UI', size=15, bold=True)  # Bold

    # Change the font style to bold for row 8
    for cell in sheet[8]:
        cell.font = font_style_bold

    # Define the alignment style for centering text
    alignment_style = Alignment(horizontal='center')

    # Define the font style with size 15
    font_style_big = Font(size=15, name="Segoe UI")

    # Center all text and apply size 15 from row 9 onwards
    for row in sheet.iter_rows(min_row=9):
        for cell in row:
            cell.alignment = alignment_style
            cell.font = font_style_big

    # Set the print area in the XLS file
    sheet.print_area = 'B1:C{}'.format(len(txt_data) + 8)  # Adjust based on the number of rows in the data
    
    # Save the modified XLS file
    xl.save(path + vocab_list_name + ".xlsm")




def main():
    # Root dir
    root_dir = "$$$ADD_YOUR_ROOT_DIR$$$/"

    # Checking language to process
    # Example languages provided
    if (len(sys.argv) < 2):
        exit()
    else:
        if (sys.argv[1] in "EUS"):
            lang_name = "Euskara/"
            vocab_list_name = "Hiztegia"
            notetype = "Basic"
        elif (sys.argv[1] in "KR"):
            lang_name = "Hangugeo/"
            vocab_list_name = "eohwi"
            notetype = "Basic (type in the answer)"
        else:
            lang_name = "Suomi/"
            vocab_list_name = "Sanasto"
            notetype = "Basic"

    anki_txt_filename = "(" + sys.argv[1] + ")" + " " + vocab_list_name + ".txt"
    lang_name += "Vocabulary/"

    # Path
    path = root_dir + lang_name

    # Anki TXT to XLS
    anki2xls(path, anki_txt_filename, vocab_list_name, notetype)


# Main
main()